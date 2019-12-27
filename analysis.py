import openpyxl,sys, datetime
import setting, os, pairs
import pandas as pd
import pair_back_test
import fileutil as ft
from dateutil.relativedelta import relativedelta

excel_file_name = 'pair_trade.xlsm'
sheet_name_history = 'Trade History'
sheet_name_open_position = 'Open Position'

def main(targetYear=None, targetMonth=None):
    print('Aalysis main start!')
    file_name = os.path.join(setting.get_master_dir(), excel_file_name)
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    sheet = workbook[sheet_name_history]
    record_list = []

    ft.clean_target_dir(os.path.join(setting.get_input_data_dir(), 'trade', 'open'))
    for i in range(4, sheet.max_row + 1, 2):
        record = TradeRecord()
        year = sheet.cell(row=i, column=27).value
        if (year is None):
            continue
        record.year = str(year)
        record.month = str(sheet.cell(row = i, column =28).value)
        record.sellCode = str(sheet.cell(row = i, column= 6).value)
        record.buyCode = str(sheet.cell(row = i + 1, column= 6).value)
        record_list.append(record)

    open_position_sheet = workbook[sheet_name_open_position]
    for i in range(4, open_position_sheet.max_row + 1, 2):
        record = TradeRecord()
        record.sellCode = str(open_position_sheet.cell(row = i, column= 6).value)
        record.buyCode = str(open_position_sheet.cell(row = i + 1, column= 6).value)
        record.kbn = 'open'
        record_list.append(record)

    for record in record_list:
        symb1=record.sellCode
        symb2=record.buyCode
        if record.sellCode > record.buyCode:
            symb1 = record.buyCode
            symb2 = record.sellCode
        file_name = symb1 + '_' + symb2 + '.csv'

        if (record.kbn != 'open'):
            target_ymd = datetime.datetime(year=int(record.year), month=int(record.month), day=1)
            ymd_edit = target_ymd + relativedelta(months=2)
            if(datetime.datetime.now() > ymd_edit):
                continue

        if (record.kbn == 'open'):
            ft.create_target_dir(os.path.join(setting.get_input_data_dir(), 'trade', record.kbn))
            csv_file_full_path = os.path.join(setting.get_input_data_dir(), 'trade', record.kbn, file_name)
        else:
            ft.create_target_dir(os.path.join(setting.get_input_data_dir(), 'trade', record.year))
            ft.create_target_dir(os.path.join(setting.get_input_data_dir(), 'trade', record.year, record.month))
            csv_file_full_path = os.path.join(setting.get_input_data_dir(), 'trade', record.year, record.month, file_name)

        cacluate_needed_data(symb1, symb2, csv_file_full_path)

    print('Aalysis main end!')

def cacluate_needed_data(symb1, symb2, csv_file_full_path):

    _pairs = pairs.create_pairs_dataframe(setting.get_input_data_dir(), symb1, symb2)
    _pairs = _pairs.sort_values('DATE', ascending=True)
    _pairs = pairs.calculate_spread_zscore(_pairs, symb1, symb2)

    if ft.is_file_exists(csv_file_full_path):
        csv_pairs = ft.read_csv(csv_file_full_path)
        csv_pairs['DATE'] = pd.to_datetime(csv_pairs['DATE'])
        csv_pairs = csv_pairs.sort_values('DATE', ascending=True)
        csv_pairs.index = csv_pairs['DATE']

        last_row_date = csv_pairs.tail(1).index
        # print ('last_row_date {0}'.format(last_row_date))

        _pairs = _pairs.combine_first(csv_pairs)
        _pairs = _pairs.loc[:,
                 ['OPEN_' + symb1, 'CLOSE_' + symb1, 'OPEN_' + symb2, 'CLOSE_' + symb2,
                  'saya_divide', 'saya_divide_mean', 'saya_divide_std', 'saya_divide_sigma',
                  'deviation_rate(%)', 'CORR_3M', 'COINT_3M', 'CORR_1Y', 'COINT_1Y']]

        _pairs = _pairs.sort_values('DATE', ascending=False)
        pair_back_test.set_corr_and_coint(_pairs, symb1, symb2, last_row_date)
        ft.write_csv(_pairs, csv_file_full_path)

    else:
        _pairs = _pairs.sort_values('DATE', ascending=False)
        pair_back_test.set_corr_and_coint(_pairs, symb1, symb2)
        ft.write_csv(_pairs, csv_file_full_path)

class TradeRecord:
    id = ""
    year=""
    month=""
    sellCode = ""
    buyCode = ""
    name = ""
    kbn = ""
    def printAll(self):
        print(self.year + self.month + self.sellCode + self.buyCode)

if __name__ == '__main__':
    args = sys.argv
    main()