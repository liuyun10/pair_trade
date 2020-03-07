import openpyxl,sys,argparse,logging
import setting, os, pairs, tradeutil
import pandas as pd
import fileutil as ft
from log_helper import LogHelper
import numpy as np
from time import strftime

excel_file_full_path = setting.get_trade_excel_full_path()
history_output_csv_file_full_path = os.path.join(setting.get_trade_history_analysis_dir(), 'history', 'trade_history.csv')
analysis_output_csv_file_dir = os.path.join(setting.get_trade_history_analysis_dir(), 'analysis')
runDateTime = strftime("%Y%m%d%H%M%S")
##################################################################################################
# Default parameters                                                                              #
##################################################################################################
# Other exit condition
EXIT_MAX_OPEN_DAYS=15
# Stop Loss
EXIT_stop_loss_rate=-0.025 # total 2.5%
# Stop Profit
EXIT_stop_profit_rate=0.025 # total 2.5%
# Trailing stop loss strategy
#None: not in use
TRAILING_STOP_LOSS_STRATEGY_NONE = None
# 1: exit when totalreturn(yesterday) > 0 and totalreturn(today) < totalreturn(yesterday)
TRAILING_STOP_LOSS_STRATEGY_1 = '1'
# 2: exit when totalreturn(yesterday) > 0.01 and totalreturn(today) < totalreturn(yesterday)
TRAILING_STOP_LOSS_STRATEGY_2 = '2'

EXIT_threshold_when_plus_sigma = None
EXIT_threshold_when_minus_sigma = None
##################################################################################################
# Define parameters                                                                              #
##################################################################################################
parser = argparse.ArgumentParser()
parser.add_argument("--trade_data_path", type=str, default="./stock_data/trade/",
                    help="Path to trade stock price data")
parser.add_argument("--log_output_dir", type=str, default="./stock_data/trade/trade_history_analysis/logs", help="Path to output log directory.")
parser.add_argument("--backtest_start_ym", type=str, default=None,
                    help="Start date YM(close) of backtest. e.g. 201911")
parser.add_argument("--backtest_end_ym", type=str, default=None,
                    help="End date YM(close) of backtest. e.g. 202111")
parser.add_argument("--trailing_stop_loss_strategy", default=TRAILING_STOP_LOSS_STRATEGY_NONE, type=str, choices=["None", "1", "2"],
                    help="Trailing stop loss strategy used.")
parser.add_argument("--exit_threshold_when plus_sigma", default=EXIT_threshold_when_plus_sigma, type=str,
                    help="Exit threshold value to be tested (in units 'number of SD from mean').")
parser.add_argument("--exit_threshold_when minux_sigma", default=EXIT_threshold_when_minus_sigma, type=str,
                    help="Exit threshold value to be tested (in units 'number of SD from mean').")
parser.add_argument("--stop_profit_limit", default=EXIT_stop_profit_rate, type=float,
                    help="Position will exit if total 2 pairs profit exceeded this loss limit.")
parser.add_argument("--stop_loss_limit", default=EXIT_stop_loss_rate, type=float,
                    help="Position will exit if total 2 pairs loss exceeded this loss limit.")
parser.add_argument("--loss_limit_max_days", default=EXIT_MAX_OPEN_DAYS, type=int,
                    help="Position will exit if loss exceeded this max openning day limit.")
parser.add_argument("--clear_all_analysis_data", default=False, type=bool,
                    help="Clear all current analysis data directories including files.")
config = parser.parse_args()

# Setup logger
LogHelper.setup(log_path='{0}/backtesting_{1}.log'.format(config.log_output_dir, runDateTime), log_level=logging.INFO)
_logger = logging.getLogger(__name__)

def main():
    # Log all paremeters
    _logger.info("Backtest parameters: {}".format(vars(config)))

    pre_action()
    output_trade_history_to_csv()
    generate_trade_result()

def pre_action():
    if (config.clear_all_analysis_data):
        _logger.info("Clear all analysis old data under {}".format(analysis_output_csv_file_dir))
        ft.clean_target_dir(analysis_output_csv_file_dir)

def generate_trade_result():

    trade_history_data = ft.read_csv(history_output_csv_file_full_path)
    trade_history_data['open_date'] = pd.to_datetime(trade_history_data['open_date'])
    outputList = pd.DataFrame(columns=record_column())

    for index, row in trade_history_data.iterrows():
        year = str(int(row.YEAR))
        month = str(int(row.MONTH))
        sellCode = str(int(row.sellCode))
        buytCode = str(int(row.buyCode))
        _logger.debug("Processing {0}/{1} {2} - {3}...{4}/{5}".format(index + 1, len(trade_history_data), sellCode, buytCode, year, month))

        trade_data = getTargetTradeData(year, month, sellCode, buytCode)
        if trade_data is None:
            continue
        search_data = trade_data.loc[trade_data['DATE'] >= row.open_date]
        result_row = generate_result(search_data, sellCode, buytCode, row)
        # print(result_row)
        outputList = outputList.append(result_row, ignore_index=True)

    outputAnalysisCsvFile(outputList)

def outputAnalysisCsvFile(dataList):
    targetDir = os.path.join(analysis_output_csv_file_dir, runDateTime)
    ft.clean_target_dir(targetDir)
    ft.write_csv_without_index(dataList, os.path.join(targetDir, 'trade_analysis_' + runDateTime + '.csv'))

    configFile = open(os.path.join(analysis_output_csv_file_dir, runDateTime, 'config_' + runDateTime + '.txt'), "w")
    configFile.write(str(vars(config)))
    configFile.close()

def generate_result(search_data, sellCode, buyCode, history_row):
    trade_start_date = history_row.open_date
    sellOpenPrice = history_row.sellOpenPrice
    sellMount = history_row.sellMount
    buyOpenPrice = history_row.buyOpenPrice
    buyMount = history_row.buyMount
    buyOtherFee = history_row.buyOtherFee
    sellOtherFee = history_row.sellOtherFee
    totalMount = sellOpenPrice * sellMount + buyOpenPrice * buyMount

    result_row = history_row.copy()

    try:
        search_data.at[trade_start_date, "DATE"]
    except KeyError  as e:
        _logger.error("No target data found. {0} - {1}...{2}".format(sellCode, buyCode, trade_start_date))

    trade_result = ""
    yesterday_return = 0

    for index, row in search_data.iterrows():
        open_days = int((index - trade_start_date) / np.timedelta64(1, 'D'))
        if (open_days == 0):
            continue
        elif(open_days > config.loss_limit_max_days):
            trade_result = 'OVER MAX DAY'

        buyTradeFee = tradeutil.get_trade_commission2(buyOpenPrice, buyMount)
        sellTradeFee = tradeutil.get_trade_commission2(sellOpenPrice, sellMount)
        buyCreditFee = tradeutil.get_credit_commission_for_buy_position(buyOpenPrice, buyMount, open_days)
        sellCreditFee = tradeutil.get_credit_commission_for_sell_position(sellOpenPrice, sellMount, open_days)

        if (len(trade_result) > 0):
            # close position when next day opening
            sellNowOpenPrice = row['OPEN_' + sellCode]
            buyNowOpenPrice = row['OPEN_' + buyCode]

            totalReturn = (buyNowOpenPrice - buyOpenPrice) * buyMount + (sellOpenPrice - sellNowOpenPrice) * sellMount
            totalReturn = totalReturn - buyTradeFee - sellTradeFee - buyCreditFee - sellCreditFee
            totalReturn = totalReturn - buyOtherFee - sellOtherFee
            totalReturnRate = totalReturn / totalMount

            result_row['close_date'] = index
            result_row['sellClosePrice'] = sellNowOpenPrice
            result_row['buyClosePrice'] = buyNowOpenPrice
            result_row['totalReturn'] = totalReturn
            result_row['totalReturnRate'] = totalReturnRate
            result_row['open_days'] = open_days
            result_row['result'] = trade_result

            result_row['buyTradeFee'] = buyTradeFee
            result_row['buyInterestFee'] = buyCreditFee
            result_row['sellTradeFee'] = sellTradeFee
            result_row['sellInterestFee'] = sellCreditFee
            return result_row

        sellNowClosePrice = row['CLOSE_' + sellCode]
        buyNowClosePrice = row['CLOSE_' + buyCode]
        now_return = (buyNowClosePrice - buyOpenPrice) * buyMount + (sellOpenPrice - sellNowClosePrice) * sellMount
        now_return = now_return - buyTradeFee - sellTradeFee - buyCreditFee - sellCreditFee

        if open_days >= config.loss_limit_max_days:
            trade_result = 'OVER MAX DAY'
        elif now_return < 0 and now_return / totalMount <= config.stop_loss_limit:
            trade_result = 'STOP LOSS OVER RATIO'
        elif now_return > 0 and now_return / totalMount >= config.stop_profit_limit:
            trade_result = 'STOP PROFIT OVER RATIO'

        if (len(trade_result) == 0):
            # check for trailing stop loss strategy
            if (config.trailing_stop_loss_strategy is not None):
                if (config.trailing_stop_loss_strategy == TRAILING_STOP_LOSS_STRATEGY_1):
                    if (yesterday_return > 0 and now_return > yesterday_return):
                        trade_result = 'TRAILING_STOP_LOSS_STRATEGY_1'
                elif (config.trailing_stop_loss_strategy == TRAILING_STOP_LOSS_STRATEGY_2):
                    pass
        yesterday_return = now_return

def getTargetTradeData(year, month, sellCode, buyCode):
    csvFile = sellCode + "_" + buyCode + '.csv'
    if (sellCode > buyCode):
        csvFile = buyCode + "_" + sellCode + '.csv'
    targetFilePath = os.path.join(setting.get_trade_dir(), year, month, csvFile)
    if not os.path.exists(targetFilePath):
        _logger.error("No target csv file exists. {0} - {1}...{2}/{3}...{4}".format(sellCode, buyCode, year, month, targetFilePath))
        return

    trade_data = ft.read_csv(targetFilePath)
    trade_data['DATE'] = pd.to_datetime(trade_data['DATE'])
    trade_data.index = trade_data['DATE']
    trade_data = trade_data.sort_index(ascending=True)
    return trade_data

def output_trade_history_to_csv():

    workbook = openpyxl.load_workbook(excel_file_full_path, data_only=True)
    sheet = workbook[setting.sheet_name_history]

    data_list = []

    for i in range(4, sheet.max_row + 1, 2):
        record = TradeHistoryRecord()
        year = sheet.cell(row=i, column=27).value
        if (year is None):
            continue

        record.year = str(year)
        record.month = str(sheet.cell(row = i, column =28).value)

        record.open_date= str(sheet.cell(row = i, column= 4).value).replace(' 00:00:00','')
        record.close_date = str(sheet.cell(row=i, column=21).value).replace(' 00:00:00','')

        record.sellCode = str(sheet.cell(row = i, column= 6).value)
        record.sellCodeName = str(sheet.cell(row=i, column=7).value)
        record.sellOpenPrice = str(sheet.cell(row = i, column= 8).value)
        record.sellMount = str(sheet.cell(row=i, column=9).value)
        record.sellClosePrice = str(sheet.cell(row=i, column=12).value)
        record.sellTradeFee = str(sheet.cell(row = i, column= 13).value)
        record.sellInterestFee = str(sheet.cell(row=i, column=14).value)
        record.sellOtherFee = str(sheet.cell(row=i, column=15).value)

        record.buyCode = str(sheet.cell(row = i + 1, column= 6).value)
        record.buyCodeName = str(sheet.cell(row= i + 1, column=7).value)
        record.buyOpenPrice = str(sheet.cell(row = i + 1, column= 8).value)
        record.buyMount = str(sheet.cell(row=i + 1, column=9).value)
        record.buyClosePrice = str(sheet.cell(row=i + 1, column=12).value)
        record.buyTradeFee = str(sheet.cell(row = i + 1, column= 13).value)
        record.buyInterestFee = str(sheet.cell(row=i + 1, column=14).value)
        record.buyOtherFee = str(sheet.cell(row=i + 1, column=15).value)

        record.totalReturn = str(sheet.cell(row=i, column=18).value)
        record.totalReturnRate = str(sheet.cell(row=i, column=19).value)
        record.open_days = str(sheet.cell(row=i, column=22).value)
        record.result = str(sheet.cell(row=i, column=23).value)

        data_list.append(convert_record_to_list(record))

    pd_data = pd.DataFrame(data_list, columns=record_column())
    ft.write_csv_without_index(pd_data, history_output_csv_file_full_path)

def convert_record_to_list(data):
    row_list = []
    row_list.append(data.year)
    row_list.append(data.month)

    row_list.append(data.open_date)
    row_list.append(data.close_date)

    row_list.append(data.sellCode)
    row_list.append(data.sellCodeName)
    row_list.append(data.sellOpenPrice)
    row_list.append(data.sellMount)
    row_list.append(data.sellClosePrice)
    row_list.append(data.sellTradeFee)
    row_list.append(data.sellInterestFee)
    row_list.append(data.sellOtherFee)

    row_list.append(data.buyCode)
    row_list.append(data.buyCodeName)
    row_list.append(data.buyOpenPrice)
    row_list.append(data.buyMount)
    row_list.append(data.buyClosePrice)
    row_list.append(data.buyTradeFee)
    row_list.append(data.buyInterestFee)
    row_list.append(data.buyOtherFee)

    row_list.append(data.totalReturn)
    row_list.append(data.totalReturnRate)
    row_list.append(data.open_days)
    row_list.append(data.result)
    return row_list

def record_column():
    columns = ['YEAR', 'MONTH', 'open_date', 'close_date',
               'sellCode', 'sellCodeName', 'sellOpenPrice', 'sellMount', 'sellClosePrice', 'sellTradeFee', 'sellInterestFee', 'sellOtherFee',
               'buyCode', 'buyCodeName', 'buyOpenPrice', 'buyMount', 'buyClosePrice', 'buyTradeFee', 'buyInterestFee', 'buyOtherFee',
               'totalReturn','totalReturnRate', 'open_days','result']
    return columns

class TradeHistoryRecord:
    id = ""
    open_date = ""
    close_date = ""
    sellCode = ""
    sellCodeName = ""
    sellOpenPrice = ""
    sellClosePrice = ""
    sellMount = ""
    sellTradeFee = ""
    sellInterestFee = ""
    sellOtherFee = ""
    buyCode = ""
    buyCodeName=""
    buyOpenPrice = ""
    buyClosePrice = ""
    buyMount=""
    buyTradeFee = ""
    buyInterestFee = ""
    buyOtherFee = ""
    name = ""
    kbn = ""
    totalReturn = ""
    totalReturnRate = ""
    open_days = ""
    year=""
    month=""
    result = ""

    def printAll(self):
        print(self.year + self.month + self.sellCode + self.buyCode)

if __name__ == '__main__':
    args = sys.argv
    main()