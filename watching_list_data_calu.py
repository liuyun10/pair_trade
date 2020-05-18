import openpyxl,sys, datetime
import setting, os, pairs
import pandas as pd
import pair_back_test
import fileutil as ft
from dateutil.relativedelta import relativedelta
import pairs as pairs_main
import tradeutil as trade_util
import fileutil as file_util
from operator import itemgetter

excel_file_name = 'pair_trade.xlsm'
sheet_name_Watching_Input = 'Watching_Input'

def main(targetYear=None, targetMonth=None):
    print('Watching List data caculate main start!')
    file_name = os.path.join(setting.get_root_dir(), excel_file_name)
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    sheet = workbook[sheet_name_Watching_Input]
    record_list = []
    symbols_corr_list = []

    ft.clean_target_dir(os.path.join(setting.get_watching_list_file_dir()))
    for i in range(4, sheet.max_row + 1, 1):

        record = WatchingRecord()
        code1 = str(sheet.cell(row=i, column=3).value)
        code2 = str(sheet.cell(row=i, column=7).value)
        if (code1 is None or code2 is None):
            continue
        record.code1 = code1
        record.code2 = code2
        record_list.append(record)

    for record in record_list:
        symb1=record.code1
        symb2=record.code2

        if (symb1 is None or symb2 is None or len(symb1) <=0 or len(symb2) <=0 or symb1 == "None" or symb2 == "None"):
            continue

        _pairs = pairs_main.create_pairs_dataframe(setting.get_input_data_dir(), symb1, symb2)
        corr_3m, corr_1y = trade_util.check_corr(_pairs, symb1, symb2)
        coint_3m, coint_1y = trade_util.check_cointegration(_pairs, symb1, symb2)
        symbols_corr_list.append([symb1, symb2, corr_3m, corr_1y, coint_3m, coint_1y])
        _pairs = _pairs.sort_values('DATE', ascending=True)
        _pairs = pairs_main.calculate_spread_zscore(_pairs, symb1, symb2)
        _pairs = _pairs.sort_values('DATE', ascending=False)
        file_util.write_csv(_pairs, os.path.join(setting.get_watching_list_file_dir(), symb1 + '_' + symb2 + '.csv'))

    corr_data = sorted(symbols_corr_list, key=itemgetter(3), reverse=True)  # sort by 3 month corr
    corr_data = pd.DataFrame(columns=['SYM_A', 'SYM_B', 'CORR_3M', 'CORR_1Y', 'COINT_3M', 'COINT_1Y'],
                             data=corr_data)
    # file_util.write_csv(corr_data, os.path.join(setting.get_result_dir(), corr_result_file_name))

    pairs_main.output_report(corr_data, False, setting.get_watching_list_file_dir(), setting.watching_corr_result_file_name)

    print('Watching List data caculate main end!')

class WatchingRecord:
    id = ""
    code1 = ""
    code2 = ""
    def printAll(self):
        print(self.id + self.code1 + self.code2)

if __name__ == '__main__':
    args = sys.argv
    main()